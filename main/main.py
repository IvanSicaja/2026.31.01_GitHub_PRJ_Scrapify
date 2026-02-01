import sys
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import date
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout,
    QWidget, QLabel, QMessageBox, QProgressBar, QTextEdit, QComboBox,
    QRadioButton, QGroupBox
)
from PyQt6.QtGui import QPalette, QColor
from PyQt6.QtCore import Qt, QThread, pyqtSignal

# ---------------------------------------------------------------------------
# SITE CONFIGURATIONS
# ---------------------------------------------------------------------------
# Each entry describes how to scrape one company's careers page.
#
# Keys:
#   url            – the careers listing page URL
#   listing_parser – callable(soup) → list of (title: str, detail_url: str)
#                    Extracts job title + absolute detail-page URL from the listing page.
#   detail_parser  – callable(detail_soup) → str
#                    Extracts the full job description text from a single job detail page.
#   headers        – optional dict of extra HTTP headers (e.g. User-Agent spoofs)
#   note           – human-readable note shown in log when the site needs special handling
# ---------------------------------------------------------------------------

HEADERS_DEFAULT = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


# ── Lever.co (shared by RIVR & Gravis Robotics) ──────────────────────────
def lever_listing_parser(soup):
    jobs = []
    for post in soup.find_all("div", class_="posting"):
        title_elem = post.find("h5")
        if not title_elem:
            continue
        title = title_elem.text.strip()
        link_elem = post.find("a", class_="posting-title")
        if not link_elem or not link_elem.get("href"):
            continue
        jobs.append((title, link_elem["href"]))
    return jobs


def lever_detail_parser(detail_soup):
    sections = detail_soup.find_all("div", class_="section page-centered")
    return "\n\n".join(s.get_text(separator="\n", strip=True) for s in sections)


# ── ANYbotics (WordPress + Workable embed) ───────────────────────────────
# The open-positions section loads via a Workable iframe/script.
# We scrape Workable's public API endpoint for ANYbotics directly.
ANYBOTICS_WORKABLE_SUBDOMAIN = "anybotics"


def anybotics_listing_parser(soup):
    """
    ANYbotics embeds Workable. We reach Workable's public JSON API instead
    of parsing the HTML (which contains no job data in static HTML).
    Returns jobs fetched via the API; detail URLs point to Workable pages.
    """
    # This is called with the original soup, but we ignore it and fetch from API.
    # The actual API call happens here because we need network access.
    api_url = f"https://api.workable.com/organization/anybotics/jobs"
    try:
        resp = requests.get(api_url, headers=HEADERS_DEFAULT, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        jobs = []
        for job in data.get("jobs", []):
            title = job.get("title", "").strip()
            url = job.get("url", "").strip()
            if title and url:
                jobs.append((title, url))
        return jobs
    except Exception:
        # Fallback: try scraping the careers page for any <a> links that
        # contain "workable" or job-like patterns
        return []


def anybotics_detail_parser(detail_soup):
    # Workable job pages put description in <div class="description">
    desc_div = detail_soup.find("div", class_="description")
    if desc_div:
        return desc_div.get_text(separator="\n", strip=True)
    # Generic fallback
    main = detail_soup.find("main") or detail_soup.find("div", class_="content")
    if main:
        return main.get_text(separator="\n", strip=True)
    return detail_soup.get_text(separator="\n", strip=True)


# ── Mimic Robotics (Framer / JS-rendered) ─────────────────────────────────
def mimic_listing_parser(soup):
    # Framer sites render jobs via JS; static HTML has no job data.
    # We attempt a fallback: look for any <a> tags whose href contains
    # common job-path patterns.
    jobs = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(kw in href.lower() for kw in ["/jobs/", "/careers/", "/job/"]):
            title = a.get_text(strip=True)
            if title and len(title) > 3:
                jobs.append((title, href if href.startswith("http") else "https://www.mimicrobotics.com" + href))
    return jobs


def mimic_detail_parser(detail_soup):
    for tag in ["main", "article"]:
        el = detail_soup.find(tag)
        if el:
            return el.get_text(separator="\n", strip=True)
    body = detail_soup.find("body")
    return body.get_text(separator="\n", strip=True) if body else ""


# ── Leica Geosystems (Hexagon careers platform) ──────────────────────────
# Their job-feed page uses a Hexagon careers widget that loads via JS.
# We attempt to hit the underlying API used by their careers platform.
def leica_listing_parser(soup):
    # Leica uses Hexagon's careers platform which loads jobs via XHR.
    # Try their known careers API endpoint pattern:
    try:
        api = "https://careers.hexagon.com/api/v1/jobs?brand=leica-geosystems&limit=100"
        resp = requests.get(api, headers=HEADERS_DEFAULT, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        jobs = []
        for job in data.get("jobs", data.get("results", [])):
            title = job.get("title", job.get("name", "")).strip()
            url = job.get("url", job.get("link", "")).strip()
            if title and url:
                jobs.append((title, url))
        return jobs
    except Exception:
        # Fallback: scan static HTML for job links
        jobs = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "job" in href.lower() and ("leica" in href.lower() or "hexagon" in href.lower()):
                title = a.get_text(strip=True)
                if title and len(title) > 3:
                    jobs.append((title, href if href.startswith("http") else "https://leica-geosystems.com" + href))
        return jobs


def leica_detail_parser(detail_soup):
    for cls in ["job-description", "description", "content"]:
        el = detail_soup.find(class_=cls)
        if el:
            return el.get_text(separator="\n", strip=True)
    main = detail_soup.find("main")
    if main:
        return main.get_text(separator="\n", strip=True)
    return ""


# ── Hexagon Robotics (onlyfy.jobs platform) ───────────────────────────────
def hexagon_robotics_listing_parser(soup):
    jobs = []
    # Jobs are in <h5> tags that contain <a> links to onlyfy.jobs
    for h5 in soup.find_all("h5"):
        a = h5.find("a", href=True)
        if a:
            title = a.get_text(strip=True)
            url = a["href"]
            if title and url:
                jobs.append((title, url if url.startswith("http") else "https://robotics.hexagon.com" + url))
    return jobs


def hexagon_robotics_detail_parser(detail_soup):
    # onlyfy.jobs renders job content via JS; try to grab whatever is available
    for cls in ["job-description", "description", "content", "oj-content"]:
        el = detail_soup.find(class_=cls)
        if el:
            return el.get_text(separator="\n", strip=True)
    # Fallback: get all meaningful text
    main = detail_soup.find("main") or detail_soup.find("body")
    if main:
        text = main.get_text(separator="\n", strip=True)
        # Filter out navigation noise
        lines = [l for l in text.splitlines() if len(l.strip()) > 20]
        return "\n".join(lines)
    return ""


# ── Flexion Robotics (Framer) ─────────────────────────────────────────────
def flexion_listing_parser(soup):
    jobs = []
    seen = set()
    for h5 in soup.find_all("h5"):
        a = h5.find("a", href=True)
        if not a:
            # Check if the h5 itself is inside an <a>
            parent_a = h5.find_parent("a", href=True)
            if parent_a:
                a = parent_a
            else:
                continue
        title = h5.get_text(strip=True)
        href = a["href"]
        # Build absolute URL
        if href.startswith("./"):
            url = "https://flexion.ai" + href[1:]  # ./careers/X → /careers/X
        elif href.startswith("/"):
            url = "https://flexion.ai" + href
        elif not href.startswith("http"):
            url = "https://flexion.ai/" + href
        else:
            url = href

        if title and url and url not in seen:
            seen.add(url)
            jobs.append((title, url))
    return jobs


def flexion_detail_parser(detail_soup):
    # Framer renders content in custom data-attributes or generic divs
    # Try common patterns first
    for cls in ["description", "job-description", "content"]:
        el = detail_soup.find(class_=cls)
        if el:
            return el.get_text(separator="\n", strip=True)
    # Fallback: grab body text, filter short lines
    body = detail_soup.find("body")
    if body:
        text = body.get_text(separator="\n", strip=True)
        lines = [l for l in text.splitlines() if len(l.strip()) > 15]
        return "\n".join(lines)
    return ""


# ── Flyability ─────────────────────────────────────────────────────────────
def flyability_listing_parser(soup):
    jobs = []
    # Flyability careers page typically lists jobs in <a> tags or card components
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(kw in href.lower() for kw in ["/career/", "/careers/", "/job/", "/position/"]):
            title = a.get_text(strip=True)
            if title and len(title) > 3 and not any(
                skip in title.lower() for skip in ["open position", "back", "apply", "see all", "view all"]
            ):
                url = href if href.startswith("http") else "https://www.flyability.com" + href
                jobs.append((title, url))
    # Deduplicate by title
    seen_titles = set()
    unique = []
    for t, u in jobs:
        if t not in seen_titles:
            seen_titles.add(t)
            unique.append((t, u))
    return unique


def flyability_detail_parser(detail_soup):
    for cls in ["job-description", "description", "content", "entry-content"]:
        el = detail_soup.find(class_=cls)
        if el:
            return el.get_text(separator="\n", strip=True)
    main = detail_soup.find("main") or detail_soup.find("article")
    if main:
        return main.get_text(separator="\n", strip=True)
    return ""


# ---------------------------------------------------------------------------
# MASTER COMPANY DATABASE
# ---------------------------------------------------------------------------
COMPANIES = {
    "RIVR": {
        "url": "https://jobs.lever.co/rivr",
        "listing_parser": lever_listing_parser,
        "detail_parser": lever_detail_parser,
        "headers": HEADERS_DEFAULT,
    },
    "ANYbotics": {
        "url": "https://www.anybotics.com/about-us/careers/",
        "listing_parser": anybotics_listing_parser,
        "detail_parser": anybotics_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Uses Workable API – static HTML has no job listings.",
    },
    "Gravis Robotics": {
        "url": "https://jobs.lever.co/gravisrobotics",
        "listing_parser": lever_listing_parser,
        "detail_parser": lever_detail_parser,
        "headers": HEADERS_DEFAULT,
    },
    "Mimic Robotics": {
        "url": "https://www.mimicrobotics.com/careers",
        "listing_parser": mimic_listing_parser,
        "detail_parser": mimic_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Framer site – jobs may be JS-rendered; static scrape may return 0 results.",
    },
    "Leica Geosystems": {
        "url": "https://leica-geosystems.com/about-us/careers/explore/job-feed",
        "listing_parser": leica_listing_parser,
        "detail_parser": leica_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Hexagon careers platform – attempts API fallback if page blocks scrape.",
    },
    "Hexagon Robotics": {
        "url": "https://robotics.hexagon.com/careers/",
        "listing_parser": hexagon_robotics_listing_parser,
        "detail_parser": hexagon_robotics_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Detail pages on onlyfy.jobs are JS-rendered; descriptions may be partial.",
    },
    "Flexion Robotics": {
        "url": "https://flexion.ai/careers",
        "listing_parser": flexion_listing_parser,
        "detail_parser": flexion_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Framer site – detail pages are JS-rendered; descriptions may be partial.",
    },
    "Flyability": {
        "url": "https://www.flyability.com/career",
        "listing_parser": flyability_listing_parser,
        "detail_parser": flyability_detail_parser,
        "headers": HEADERS_DEFAULT,
        "note": "Detail page structure may vary; uses generic fallback parser.",
    },
}

# Companies without a provided URL (placeholders – skipped during scrape)
COMPANIES_NO_URL = ["Hexagon AB", "Flink Robotics"]


# ---------------------------------------------------------------------------
# WORKER THREAD
# ---------------------------------------------------------------------------
class Worker(QThread):
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    log = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, selected_companies: list):
        super().__init__()
        self.selected_companies = selected_companies

    def run(self):
        try:
            today_date = date.today().strftime("%m/%d/%Y")
            all_jobs = []          # (date, company, title, description)
            total_companies = len(self.selected_companies)

            for idx, company_name in enumerate(self.selected_companies):
                cfg = COMPANIES.get(company_name)
                if not cfg:
                    self.log.emit(f"⚠ {company_name} — no configuration, skipping.")
                    continue

                self.log.emit(f"\n── {company_name} ──")
                if cfg.get("note"):
                    self.log.emit(f"  ℹ {cfg['note']}")

                base_progress = int((idx / total_companies) * 80)
                company_progress_range = int(80 / total_companies)

                # ── 1. Fetch listing page ────────────────────────────────
                self.status.emit(f"[{company_name}] Fetching listings page…")
                try:
                    response = requests.get(
                        cfg["url"],
                        headers=cfg.get("headers", HEADERS_DEFAULT),
                        timeout=20
                    )
                    response.raise_for_status()
                except Exception as e:
                    self.log.emit(f"  ✗ Failed to fetch listing page: {e}")
                    continue

                self.progress.emit(base_progress + int(company_progress_range * 0.15))

                # ── 2. Parse listing page → job list ─────────────────────
                soup = BeautifulSoup(response.text, "html.parser")
                try:
                    job_list = cfg["listing_parser"](soup)
                except Exception as e:
                    self.log.emit(f"  ✗ Listing parser error: {e}")
                    continue

                num_jobs = len(job_list)
                if num_jobs == 0:
                    self.log.emit(f"  ⚠ No job postings found (page may require JS rendering).")
                    self.progress.emit(base_progress + company_progress_range)
                    continue

                self.log.emit(f"  Found {num_jobs} posting(s)")
                self.progress.emit(base_progress + int(company_progress_range * 0.25))

                # ── 3. Fetch each job detail page ────────────────────────
                detail_step = (company_progress_range * 0.60) / num_jobs if num_jobs else 0

                for i, (title, detail_url) in enumerate(job_list, 1):
                    self.status.emit(f"[{company_name}] {i}/{num_jobs}: {title}")
                    try:
                        detail_resp = requests.get(
                            detail_url,
                            headers=cfg.get("headers", HEADERS_DEFAULT),
                            timeout=15
                        )
                        detail_resp.raise_for_status()
                    except Exception as e:
                        self.log.emit(f"  ✗ Skipped (fetch error): {title} — {e}")
                        self.progress.emit(int(base_progress + 0.25 * company_progress_range + i * detail_step))
                        continue

                    detail_soup = BeautifulSoup(detail_resp.text, "html.parser")
                    try:
                        description = cfg["detail_parser"](detail_soup)
                    except Exception as e:
                        self.log.emit(f"  ✗ Skipped (parse error): {title} — {e}")
                        self.progress.emit(int(base_progress + 0.25 * company_progress_range + i * detail_step))
                        continue

                    if description.strip():
                        all_jobs.append((today_date, company_name, title, description.strip()))
                        self.log.emit(f"  ✔ {title}")
                    else:
                        self.log.emit(f"  ✗ No description: {title}")

                    self.progress.emit(int(base_progress + 0.25 * company_progress_range + i * detail_step))
                    time.sleep(0.4)  # polite delay between requests

                self.progress.emit(base_progress + company_progress_range)

            # ── 4. Write to Excel ────────────────────────────────────────
            if not all_jobs:
                self.finished.emit(True, "Scraping complete — no job descriptions were collected.")
                return

            self.status.emit("Preparing Excel file…")
            file_name = "Scrapify.xlsx"
            sheet_name = "Sheet1"
            try:
                wb = openpyxl.load_workbook(file_name)
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                ws.append(["Date", "Company", "Role", "Role description"])

            ws = wb[sheet_name]

            self.status.emit("Inserting jobs…")
            insert_step = 15 / len(all_jobs) if all_jobs else 0
            for j, job in enumerate(reversed(all_jobs), 1):
                ws.insert_rows(2)
                try:
                    ws.cell(row=2, column=1, value=job[0])
                    ws.cell(row=2, column=2, value=job[1])
                    ws.cell(row=2, column=3, value=job[2])
                    ws.cell(row=2, column=4, value=job[3])
                except IllegalCharacterError:
                    cleaned = "".join(c for c in job[3] if c.isprintable())
                    ws.cell(row=2, column=4, value=cleaned)
                self.progress.emit(80 + int(j * insert_step))

            self.status.emit("Saving…")
            wb.save(file_name)
            self.progress.emit(100)
            self.status.emit("Done")
            self.finished.emit(True, f"Added {len(all_jobs)} job(s) to {file_name}.")

        except Exception as e:
            self.finished.emit(False, f"An error occurred: {e}")


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------
class ScraperApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Scrapify — Job Scraper")
        self.setFixedSize(560, 520)

        self.setStyleSheet("""
            QMainWindow { background-color: #1e1e1e; }
            QLabel {
                color: #e0e0e0;
                font-size: 14px;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            }
            QGroupBox {
                color: #aaa;
                font-size: 13px;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                border: 1px solid #333;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
            }
            QRadioButton {
                color: #ccc;
                font-size: 13px;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                spacing: 6px;
            }
            QRadioButton::indicator {
                width: 16px;
                height: 16px;
                border-radius: 8px;
                border: 2px solid #555;
                background: #2a2a2a;
            }
            QRadioButton::indicator:checked {
                background: #0a84ff;
                border-color: #0a84ff;
            }
            QComboBox {
                background-color: #2a2a2a;
                color: #e0e0e0;
                border: 1px solid #444;
                border-radius: 8px;
                padding: 6px 10px;
                font-size: 13px;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            }
            QComboBox:disabled {
                background-color: #1e1e1e;
                color: #555;
                border-color: #333;
            }
            QComboBox::drop-down { border: none; }
            QComboBox::down-arrow { image: none; border: none; }
            QComboBox QAbstractItemView {
                background-color: #2a2a2a;
                color: #e0e0e0;
                border-radius: 6px;
                selection-background-color: #0a84ff;
            }
            QProgressBar {
                height: 14px;
                border-radius: 7px;
                background: #2a2a2a;
                text-align: center;
                color: #aaa;
                font-size: 11px;
            }
            QProgressBar::chunk {
                background-color: #0a84ff;
                border-radius: 7px;
            }
            QTextEdit {
                background-color: #121212;
                color: #d0d0d0;
                border: none;
                border-radius: 10px;
                padding: 10px;
                font-family: SF Mono, Menlo, Consolas, monospace;
                font-size: 12px;
            }
            QPushButton {
                background-color: #0a84ff;
                color: white;
                border: none;
                border-radius: 12px;
                padding: 12px;
                font-size: 15px;
                font-weight: 600;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            }
            QPushButton:hover { background-color: #0070e0; }
            QPushButton:pressed { background-color: #0060c0; }
            QPushButton:disabled { background-color: #444; color: #888; }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(10)

        # ── Title ─────────────────────────────────────────────────────
        title_label = QLabel("Scrapify — Job Scraper")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: 600; color: #fff;")
        layout.addWidget(title_label)

        # ── Mode selection group ──────────────────────────────────────
        mode_group = QGroupBox("Scrape target")
        mode_layout = QVBoxLayout()
        mode_layout.setContentsMargins(12, 6, 12, 8)
        mode_layout.setSpacing(6)

        self.radio_all = QRadioButton("All companies")
        self.radio_all.setChecked(True)
        self.radio_all.toggled.connect(self._on_mode_changed)
        mode_layout.addWidget(self.radio_all)

        single_row = QHBoxLayout()
        single_row.setSpacing(8)
        self.radio_single = QRadioButton("One company")
        self.radio_single.toggled.connect(self._on_mode_changed)
        single_row.addWidget(self.radio_single)

        self.company_combo = QComboBox()
        self.company_combo.setFixedHeight(30)
        self.company_combo.setEnabled(False)
        # Populate with all companies that have URLs
        for name in COMPANIES:
            self.company_combo.addItem(name)
        # Add placeholder entries for companies without URLs (disabled)
        for name in COMPANIES_NO_URL:
            self.company_combo.addItem(f"{name} (no URL)")
        single_row.addWidget(self.company_combo, stretch=1)

        mode_layout.addLayout(single_row)
        mode_group.setLayout(mode_layout)
        layout.addWidget(mode_group)

        # ── Status ────────────────────────────────────────────────────
        self.status_label = QLabel("Ready")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("color: #888; font-size: 12px;")
        layout.addWidget(self.status_label)

        # ── Progress bar ──────────────────────────────────────────────
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # ── Log area ──────────────────────────────────────────────────
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text, stretch=1)

        # ── Run button ────────────────────────────────────────────────
        self.scrape_button = QPushButton("Run Scraper")
        self.scrape_button.setFixedHeight(46)
        self.scrape_button.clicked.connect(self.start_scraper)
        layout.addWidget(self.scrape_button)

        central = QWidget()
        central.setLayout(layout)
        self.setCentralWidget(central)

    # ── Slots ───────────────────────────────────────────────────────────
    def _on_mode_changed(self):
        single_selected = self.radio_single.isChecked()
        self.company_combo.setEnabled(single_selected)

    def _get_selected_companies(self) -> list:
        if self.radio_all.isChecked():
            return list(COMPANIES.keys())
        else:
            chosen = self.company_combo.currentText()
            # If user somehow picked a placeholder without URL, warn
            if chosen in COMPANIES_NO_URL or chosen.endswith("(no URL)"):
                return []
            return [chosen]

    def start_scraper(self):
        companies = self._get_selected_companies()
        if not companies:
            QMessageBox.warning(
                self, "Scrapify",
                "The selected company has no careers URL configured.\n"
                "Please choose a different company or select 'All companies'."
            )
            return

        self.scrape_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting…")
        self.log_text.clear()

        self.worker = Worker(companies)
        self.worker.progress.connect(self.update_progress)
        self.worker.status.connect(self.update_status)
        self.worker.log.connect(self.append_log)
        self.worker.finished.connect(self.scraper_finished)
        self.worker.start()

    def update_progress(self, value: int):
        self.progress_bar.setValue(value)

    def update_status(self, text: str):
        self.status_label.setText(text)

    def append_log(self, text: str):
        self.log_text.append(text)
        self.log_text.ensureCursorVisible()

    def scraper_finished(self, success: bool, message: str):
        self.progress_bar.setValue(100 if success else self.progress_bar.value())
        self.scrape_button.setEnabled(True)
        self.progress_bar.setVisible(False)

        if success:
            QMessageBox.information(self, "Scrapify", message)
            self.status_label.setText("Done")
        else:
            QMessageBox.critical(self, "Error", message)
            self.status_label.setText("Failed")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ScraperApp()
    window.show()
    sys.exit(app.exec())